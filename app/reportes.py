"""Módulo: ReportesNotas"""

import logging
class ReportesNotas:
    """Generador de reportes para notas"""
    
    def __init__(self, db_manager):
        self.db = db_manager
    
    def reporte_mensual(self, mes, año):
        """Generar reporte mensual"""
        query = """
        SELECT 
            tipo,
            COUNT(*) as cantidad,
            SUM(valor_total) as total_valor,
            AVG(valor_total) as promedio_valor
        FROM notas_credito_debito 
        WHERE MONTH(fecha_emision) = %s AND YEAR(fecha_emision) = %s
        GROUP BY tipo
        """
        
        return self.db.fetch_all(query, (mes, año))
    
    def reporte_por_concepto(self, fecha_inicio, fecha_fin):
        """Reporte agrupado por concepto"""
        query = """
        SELECT 
            codigo_concepto,
            tipo,
            COUNT(*) as cantidad,
            SUM(valor_total) as total_valor
        FROM notas_credito_debito 
        WHERE fecha_emision BETWEEN %s AND %s
        GROUP BY codigo_concepto, tipo
        ORDER BY total_valor DESC
        """
        
        return self.db.fetch_all(query, (fecha_inicio, fecha_fin))
    
    def exportar_excel(self, datos, filename):
        """Exportar datos a Excel (requiere openpyxl)"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Notas"
            
            # Headers
            if datos:
                headers = list(datos[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header.title())
                
                # Datos
                for row, record in enumerate(datos, 2):
                    for col, value in enumerate(record.values(), 1):
                        ws.cell(row=row, column=col, value=value)
            
            wb.save(filename)
            return True
            
        except ImportError:
            logger.error("openpyxl no está instalado")
            return False
        except Exception as e:
            logger.error(f"Error exportando Excel: {e}")
            return False
